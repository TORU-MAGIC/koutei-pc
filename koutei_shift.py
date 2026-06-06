# -*- coding: utf-8 -*-
"""
工程表 2日前倒しツール
=======================
毎週の「工程表」Excel(.xlsx)を読み込み、各仕事を「2営業日前」に並べ替えた
工程表を出力します。

仕様（サンプル工程表に基づく）
- 1ブロック = 1週間。 1行目「工程表」見出し / 2行目 日付(月〜土) / 以降が仕事。
- 列A〜F = 月・火・水・木・金・土。 日付は左上セルだけが実値で、他は数式
  (=A2+1, =A2+7 …) で自動計算される自己増殖型。書式は m/d(aaa)。
- 並べ替え方式: A〜E列(月〜金)の仕事を「2営業日前」の列へ移動する。
  週をまたいで詰める:  水→月 / 木→火 / 金→水 / (翌週)月→今週木 / 火→今週金。
  ＝先頭の週から見ると  月→前週木 / 火→前週金。
  F列(土＝遅延品・特記・色凡例の欄)は動かさない(その週のまま)。
  各セルの値・塗り色・フォント・配置はそのまま一緒に移動する(罫線=枠は固定)。
- はみ出し処理(先頭週の月・火):
    prepend … 一番上に前週ブロックを足して全部残す(既定/抽出漏れゼロ)
    clamp   … 先頭週の月曜列の空きにまとめて残す
    drop    … 載せない(前週は既に過去のため)
- セル内日付テキスト: keep(そのまま/既定) または shift(完成日らしき日付も2営業日前に・試験的)
- 出力: newfile(別名 _2日前倒し.xlsx /既定) または overwrite(上書き)

依存: openpyxl のみ(標準の tkinter で GUI)。オフライン動作。
"""

import os
import re
import sys
import copy
import datetime
import traceback

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


# ----------------------------------------------------------------------------
# 設定
# ----------------------------------------------------------------------------
WORK_COLS = [1, 2, 3, 4, 5]   # A〜E = 月〜金 (並べ替え対象)
SPECIAL_COL = 6               # F = 土/特記・凡例 (動かさない)
N_WORK = len(WORK_COLS)       # 1週あたりの稼働日数 = 5
SHIFT_DAYS = 2                # 2営業日前倒し
TITLE_NORM = "工程表"          # 見出し(空白除去後)


# ----------------------------------------------------------------------------
# スタイルのコピー用ヘルパ
# ----------------------------------------------------------------------------
def grab(cell):
    """セルの中身と書式(罫線以外)を取り出して持ち運べる形にする。"""
    return {
        "value": cell.value,
        "font": copy.copy(cell.font),
        "fill": copy.copy(cell.fill),
        "alignment": copy.copy(cell.alignment),
        "number_format": cell.number_format,
    }


def put(cell, data):
    """grab() で取り出した中身と書式をセルへ書き込む(罫線は触らない)。"""
    cell.value = data["value"]
    cell.font = copy.copy(data["font"])
    cell.fill = copy.copy(data["fill"])
    cell.alignment = copy.copy(data["alignment"])
    cell.number_format = data["number_format"]


def blank_style():
    """空セル用の標準スタイル(塗りなし)。"""
    return {
        "value": None,
        "font": Font(name="ＭＳ Ｐゴシック", size=14, bold=True),
        "fill": PatternFill(fill_type=None),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "number_format": "General",
    }


def norm(v):
    if v is None:
        return ""
    return str(v).replace("　", "").replace(" ", "").strip()


# ----------------------------------------------------------------------------
# ブロック検出
# ----------------------------------------------------------------------------
def detect_blocks(ws):
    """「工程表」見出し行を探して、週ブロックの一覧を返す。"""
    title_rows = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if norm(ws.cell(row=r, column=c).value) == TITLE_NORM:
                title_rows.append(r)
                break
    if not title_rows:
        raise ValueError("「工程表」の見出しが見つかりませんでした。ひな型が想定と異なる可能性があります。")

    if len(title_rows) >= 2:
        stride = title_rows[1] - title_rows[0]
    else:
        stride = ws.max_row - title_rows[0] + 1
    height = stride - 3  # title + date + 内容(height) + 余白1
    if height < 1:
        raise ValueError("週ブロックの行数を正しく認識できませんでした。")

    blocks = []
    for t in title_rows:
        blocks.append({
            "title_row": t,
            "date_row": t + 1,
            "content_start": t + 2,
            "title_col": _title_col(ws, t),
        })
    return blocks, stride, height


def _title_col(ws, r):
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(row=r, column=c).value) == TITLE_NORM:
            return c
    return 3


# ----------------------------------------------------------------------------
# ひな型(枠・行高・見出し/日付/余白行のスタイル)の取り込み
# ----------------------------------------------------------------------------
def capture_template(ws, first_title_row, stride):
    """先頭ブロックの 1ブロック分のひな型を取り込む。
    返り値: offset(0..stride-1) ごとの行高と、列ごとの罫線/日付・見出し・余白スタイル。"""
    tpl = {"row_height": {}, "border": {}, "title": {}, "date": {}, "gap": {}}
    for off in range(stride):
        r = first_title_row + off
        tpl["row_height"][off] = ws.row_dimensions[r].height
        for c in range(1, SPECIAL_COL + 1):
            cell = ws.cell(row=r, column=c)
            tpl["border"][(off, c)] = copy.copy(cell.border)
            if off == 0:      # 見出し行
                tpl["title"][c] = grab(cell)
            elif off == 1:    # 日付行
                tpl["date"][c] = grab(cell)
            elif off >= 2 + 0 and off >= stride - 1:  # 余白行(最後)
                pass
        # 余白行(=各ブロック末尾の空行)スタイル
        if off == stride - 1:
            for c in range(1, SPECIAL_COL + 1):
                tpl["gap"][c] = grab(ws.cell(row=r, column=c))
    return tpl


# ----------------------------------------------------------------------------
# 仕事内容の読み取り
# ----------------------------------------------------------------------------
def read_content(ws, blocks, height):
    """work[(bi,c)] = [k行ぶんの grab結果] と fcol[bi] = [k行ぶん] を返す。"""
    work = {}
    fcol = {}
    for bi, b in enumerate(blocks):
        cs = b["content_start"]
        for c in WORK_COLS:
            strip = []
            for k in range(height):
                strip.append(grab(ws.cell(row=cs + k, column=c)))
            work[(bi, c)] = strip
        fstrip = []
        for k in range(height):
            fstrip.append(grab(ws.cell(row=cs + k, column=SPECIAL_COL)))
        fcol[bi] = fstrip
    return work, fcol


# ----------------------------------------------------------------------------
# セル内の完成日テキストを2営業日前に書き換える(試験的)
# ----------------------------------------------------------------------------
_DATE_TOKEN = re.compile(r"(?<![0-9(（])(\d{1,2})/(\d{1,2})(?![0-9)）/])")


def add_business_days(d, delta):
    """営業日(月〜金)ベースで delta 日ずらす。delta は負で前倒し。"""
    step = 1 if delta > 0 else -1
    remaining = abs(delta)
    cur = d
    while remaining > 0:
        cur += datetime.timedelta(days=step)
        if cur.weekday() < 5:   # 0=月 .. 4=金
            remaining -= 1
    return cur


def shift_dates_in_text(text, ref_year):
    """文中の『M/D』表記(かっこ内の受注日は除く)を2営業日前にする。"""
    if not isinstance(text, str):
        return text

    def repl(m):
        mm, dd = int(m.group(1)), int(m.group(2))
        if not (1 <= mm <= 12 and 1 <= dd <= 31):
            return m.group(0)
        try:
            d = datetime.date(ref_year, mm, dd)
        except ValueError:
            return m.group(0)
        nd = add_business_days(d, -SHIFT_DAYS)
        return "%d/%d" % (nd.month, nd.day)

    return _DATE_TOKEN.sub(repl, text)


# ----------------------------------------------------------------------------
# 並べ替えの割り付け計算
# ----------------------------------------------------------------------------
def build_mapping(nb, spillover):
    """出力ブロック数と、出力(ob,c)←元(bi,c) の対応、F列の対応、はみ出し分を返す。

    元の稼働列フラットindex i = bi*5 + (c-1)  (時系列: 上の週ほど早い)
    2営業日前  →  i-2 へ移動。
    """
    work_map = {}     # (ob, c_out) -> (bi, c_src)
    extra = []        # clamp用: 月曜にまとめる (bi, c_src) のリスト(時系列順)
    fmap = {}         # ob -> bi (F列の供給元)  なければ未登録

    if spillover == "prepend":
        out_nb = nb + 1
        # 元 i は出力フラット (i+5) に居る → 2前で (i+3)
        for bi in range(nb):
            for c in WORK_COLS:
                i = bi * N_WORK + (c - 1)
                outflat = i + N_WORK - SHIFT_DAYS  # = i+3
                ob = outflat // N_WORK
                c_out = WORK_COLS[outflat % N_WORK]
                work_map[(ob, c_out)] = (bi, c)
        for bi in range(nb):
            fmap[bi + 1] = bi          # 既存ブロックは1つ下へ
        # ob=0 のF列は供給元なし(空)
    else:
        # drop / clamp : ブロック数は据え置き
        out_nb = nb
        for bi in range(nb):
            for c in WORK_COLS:
                i = bi * N_WORK + (c - 1)
                outflat = i - SHIFT_DAYS
                if outflat < 0:
                    extra.append((bi, c))      # 先頭からはみ出す分
                    continue
                ob = outflat // N_WORK
                c_out = WORK_COLS[outflat % N_WORK]
                work_map[(ob, c_out)] = (bi, c)
        for bi in range(nb):
            fmap[bi] = bi
        if spillover == "drop":
            extra = []                          # 捨てる

    return out_nb, work_map, fmap, extra


# ----------------------------------------------------------------------------
# 出力の書き込み
# ----------------------------------------------------------------------------
def write_output(ws, tpl, work, fcol, mapping, blocks, stride, height,
                 anchor_date, rewrite_dates, ref_year, orig_max_row):
    out_nb, work_map, fmap, extra = mapping

    def set_cell(r, c, off, data):
        cell = ws.cell(row=r, column=c)
        put(cell, data)
        cell.border = copy.copy(tpl["border"][(off, c)])

    for ob in range(out_nb):
        base = ob * stride
        # 行高
        for off in range(stride):
            h = tpl["row_height"].get(off)
            if h is not None:
                ws.row_dimensions[base + off + 1].height = h

        # --- 見出し行 (off=0) ---
        for c in range(1, SPECIAL_COL + 1):
            set_cell(base + 1, c, 0, tpl["title"][c])

        # --- 日付行 (off=1) ---
        # 数式(=A2+1 等)ではなく「実際の日付の値」を書き込む。
        # 数式だと計算結果が保存されず、再計算しないビューア(タブレットのExcel/
        # プレビュー等)で日付が空欄に見えてしまうため。書式 m/d(aaa) は維持。
        date_row = base + 2
        for c in range(1, SPECIAL_COL + 1):
            data = copy.deepcopy(tpl["date"][c])
            data["value"] = anchor_date + datetime.timedelta(days=7 * ob + (c - 1))
            set_cell(date_row, c, 1, data)

        # --- 内容行 (off=2 .. 2+height-1) ---  行 = base + off + 1
        for k in range(height):
            off = 2 + k
            r = base + off + 1
            for c in WORK_COLS:
                src = work_map.get((ob, c))
                if src is None:
                    set_cell(r, c, off, blank_style())
                else:
                    data = copy.deepcopy(work[src][k])
                    if rewrite_dates:
                        data["value"] = shift_dates_in_text(data["value"], ref_year)
                    set_cell(r, c, off, data)
            # F列(特記/凡例)
            bi_src = fmap.get(ob)
            if bi_src is None:
                set_cell(r, SPECIAL_COL, off, blank_style())
            else:
                set_cell(r, SPECIAL_COL, off, copy.deepcopy(fcol[bi_src][k]))

        # --- 余白行(末尾) ---  行 = base + off + 1
        for off in range(2 + height, stride):
            r = base + off + 1
            for c in range(1, SPECIAL_COL + 1):
                set_cell(r, c, off, copy.deepcopy(tpl["gap"][c]))

    # clamp: はみ出し分を「先頭週(出力ob=0)の月曜列」の空きへ追記
    truncated = 0
    if extra:
        # extra は時系列順(古い順)。月曜列(c=1)の空き行を上から探して詰める。
        col = WORK_COLS[0]
        base = 0
        # 既に埋まっている行を把握 (内容行 = base + (2+k) + 1)
        free_ks = []
        for k in range(height):
            cell = ws.cell(row=base + 2 + k + 1, column=col)
            if cell.value in (None, ""):
                free_ks.append(k)
        # 追記対象(空でないセルだけ)を展開
        items = []
        for (bi, c) in extra:
            for k in range(height):
                d = work[(bi, c)][k]
                if d["value"] not in (None, ""):
                    items.append(d)
        for d in items:
            if not free_ks:
                truncated += 1
                continue
            k = free_ks.pop(0)
            off = 2 + k
            r = base + off + 1
            data = copy.deepcopy(d)
            if rewrite_dates:
                data["value"] = shift_dates_in_text(data["value"], ref_year)
            set_cell(r, col, off, data)

    # 出力が元より短い場合、余分な行をクリア
    out_rows = out_nb * stride
    if orig_max_row > out_rows:
        for r in range(out_rows + 1, orig_max_row + 1):
            for c in range(1, SPECIAL_COL + 1):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.fill = PatternFill(fill_type=None)

    return truncated


# ----------------------------------------------------------------------------
# 仕事数の数え上げ(抽出漏れチェック用)
# ----------------------------------------------------------------------------
def count_jobs_work(work):
    n = 0
    for strip in work.values():
        for d in strip:
            if d["value"] not in (None, ""):
                n += 1
    return n


def count_jobs_ws(ws, blocks, height, cols):
    n = 0
    for b in blocks:
        cs = b["content_start"]
        for c in cols:
            for k in range(height):
                if ws.cell(row=cs + k, column=c).value not in (None, ""):
                    n += 1
    return n


# ----------------------------------------------------------------------------
# 中核: ワークブックを2日前倒しに変換(ファイル入出力に依存しない)
# ----------------------------------------------------------------------------
def transform_workbook(wb, spillover="prepend", rewrite_dates=False, log=print):
    """開いた openpyxl ワークブックを、その場で2日前倒しに並べ替える。
    返り値: 集計(dict)。"""
    ws = wb.worksheets[0]
    orig_max_row = ws.max_row

    blocks, stride, height = detect_blocks(ws)
    nb = len(blocks)
    log("週ブロック数 = %d / 1ブロック=%d行(内容%d行)" % (nb, stride, height))

    # 先頭ブロックの日付(実値)を取得 → 前週ブロックの基準日に使う
    first_date_cell = ws.cell(row=blocks[0]["date_row"], column=1)
    first_date = first_date_cell.value
    if isinstance(first_date, datetime.datetime):
        first_date = first_date
    elif isinstance(first_date, datetime.date):
        first_date = datetime.datetime(first_date.year, first_date.month, first_date.day)
    else:
        # 数式などで実値が取れない場合のフォールバック
        log("注意: 先頭の日付が実値で取得できませんでした。今日基準で計算します。")
        first_date = datetime.datetime.now()
    ref_year = first_date.year

    # ひな型・内容を取り込み(書き換え前に)
    tpl = capture_template(ws, blocks[0]["title_row"], stride)
    work, fcol = read_content(ws, blocks, height)

    jobs_in_work = count_jobs_work(work)      # A〜E列の仕事数
    jobs_in_f = count_jobs_ws(ws, blocks, height, [SPECIAL_COL])

    # 割り付け
    mapping = build_mapping(nb, spillover)
    out_nb, work_map, fmap, extra = mapping

    # 前週ブロックの基準日(月曜) = 先頭週の月曜 - 7日
    if spillover == "prepend":
        anchor_date = first_date - datetime.timedelta(days=7)
    else:
        anchor_date = first_date

    truncated = write_output(ws, tpl, work, fcol, mapping, blocks, stride, height,
                             anchor_date, rewrite_dates, ref_year, orig_max_row)

    # 出力後の仕事数(抽出漏れチェック)
    out_blocks, _, _ = detect_blocks(ws)
    jobs_out_work = count_jobs_ws(ws, out_blocks, height, WORK_COLS)
    jobs_out_f = count_jobs_ws(ws, out_blocks, height, [SPECIAL_COL])

    return {
        "blocks_in": nb,
        "blocks_out": out_nb,
        "jobs_in_work": jobs_in_work,
        "jobs_out_work": jobs_out_work,
        "jobs_in_f": jobs_in_f,
        "jobs_out_f": jobs_out_f,
        "truncated": truncated,
        "spillover": spillover,
        "rewrite_dates": rewrite_dates,
    }


def process_bytes(data, spillover="prepend", rewrite_dates=False, log=print):
    """xlsxのバイト列を受け取り、2日前倒し後のバイト列と集計を返す。"""
    import io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(data), data_only=False)
    summary = transform_workbook(wb, spillover=spillover,
                                 rewrite_dates=rewrite_dates, log=log)
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), summary


def process(input_path, spillover="prepend", rewrite_dates=False,
            output_mode="newfile", log=print):
    """ファイルパス版(CLI/旧tkinter用)。読み込んで変換し、保存する。"""
    log("読み込み: %s" % input_path)
    wb = openpyxl.load_workbook(input_path, data_only=False)
    summary = transform_workbook(wb, spillover=spillover,
                                 rewrite_dates=rewrite_dates, log=log)
    if output_mode == "overwrite":
        out_path = input_path
    else:
        root, ext = os.path.splitext(input_path)
        out_path = root + "_2日前倒し" + ext
    wb.save(out_path)
    log("保存: %s" % out_path)
    summary["input"] = input_path
    summary["output"] = out_path
    return summary


# ----------------------------------------------------------------------------
# GUI
# ----------------------------------------------------------------------------
def run_gui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, scrolledtext

    root = tk.Tk()
    root.title("工程表 2日前倒しツール")
    root.geometry("760x560")

    pad = {"padx": 8, "pady": 4}

    # ファイル選択
    frm_file = ttk.LabelFrame(root, text="① 工程表ファイル(.xlsx)を選ぶ")
    frm_file.pack(fill="x", **pad)
    var_path = tk.StringVar()
    ent = ttk.Entry(frm_file, textvariable=var_path)
    ent.pack(side="left", fill="x", expand=True, padx=6, pady=6)

    def pick():
        p = filedialog.askopenfilename(
            title="工程表を選択",
            filetypes=[("Excel ブック", "*.xlsx"), ("すべて", "*.*")])
        if p:
            var_path.set(p)
    ttk.Button(frm_file, text="参照…", command=pick).pack(side="left", padx=6)

    # オプション
    frm_opt = ttk.LabelFrame(root, text="② オプション")
    frm_opt.pack(fill="x", **pad)

    ttk.Label(frm_opt, text="方式:  A〜E列(月〜金)の仕事を2営業日前へ移動／F列(土・特記)はそのまま",
              foreground="#555").grid(row=0, column=0, columnspan=3, sticky="w", padx=6, pady=(6, 2))

    ttk.Label(frm_opt, text="先頭週からはみ出す仕事:").grid(row=1, column=0, sticky="w", padx=6, pady=2)
    var_spill = tk.StringVar(value="prepend")
    ttk.Radiobutton(frm_opt, text="前週ブロックを追加して全部残す(推奨)", variable=var_spill, value="prepend").grid(row=1, column=1, sticky="w")
    ttk.Radiobutton(frm_opt, text="先頭週の月曜にまとめる", variable=var_spill, value="clamp").grid(row=2, column=1, sticky="w")
    ttk.Radiobutton(frm_opt, text="載せない(前週は過去)", variable=var_spill, value="drop").grid(row=3, column=1, sticky="w")

    ttk.Label(frm_opt, text="セル内の日付:").grid(row=4, column=0, sticky="w", padx=6, pady=2)
    var_rewrite = tk.BooleanVar(value=False)
    ttk.Radiobutton(frm_opt, text="文字はそのまま(推奨)", variable=var_rewrite, value=False).grid(row=4, column=1, sticky="w")
    ttk.Radiobutton(frm_opt, text="完成日も2営業日前に書き換える(試験的)", variable=var_rewrite, value=True).grid(row=5, column=1, sticky="w")

    ttk.Label(frm_opt, text="出力:").grid(row=6, column=0, sticky="w", padx=6, pady=2)
    var_out = tk.StringVar(value="newfile")
    ttk.Radiobutton(frm_opt, text="別名で保存 (_2日前倒し)(推奨)", variable=var_out, value="newfile").grid(row=6, column=1, sticky="w")
    ttk.Radiobutton(frm_opt, text="同じファイルに上書き", variable=var_out, value="overwrite").grid(row=7, column=1, sticky="w")

    # ログ
    frm_log = ttk.LabelFrame(root, text="③ 実行ログ")
    frm_log.pack(fill="both", expand=True, **pad)
    txt = scrolledtext.ScrolledText(frm_log, height=10)
    txt.pack(fill="both", expand=True, padx=6, pady=6)

    def log(msg):
        txt.insert("end", str(msg) + "\n")
        txt.see("end")
        root.update_idletasks()

    def do_run():
        p = var_path.get().strip().strip('"')
        if not p:
            messagebox.showwarning("確認", "工程表ファイルを選んでください。")
            return
        if not os.path.isfile(p):
            messagebox.showerror("エラー", "ファイルが見つかりません:\n%s" % p)
            return
        if var_out.get() == "overwrite":
            if not messagebox.askyesno("確認", "元のファイルに上書きします。よろしいですか？\n(元データは戻せません)"):
                return
        txt.delete("1.0", "end")
        btn.config(state="disabled")
        try:
            s = process(p, spillover=var_spill.get(),
                        rewrite_dates=var_rewrite.get(),
                        output_mode=var_out.get(), log=log)
            log("―" * 30)
            log("週ブロック: %d → %d" % (s["blocks_in"], s["blocks_out"]))
            log("A〜E列の仕事数: 入力 %d → 出力 %d" % (s["jobs_in_work"], s["jobs_out_work"]))
            log("F列(特記/凡例): 入力 %d → 出力 %d" % (s["jobs_in_f"], s["jobs_out_f"]))
            ok_msg = "完了しました。\n\n出力: %s" % s["output"]
            if s["spillover"] == "prepend":
                if s["jobs_out_work"] != s["jobs_in_work"]:
                    log("⚠ 注意: A〜E列の仕事数が一致しません。確認してください。")
                    ok_msg += "\n\n⚠ 仕事数が一致しません。ログを確認してください。"
                else:
                    log("✓ 抽出漏れなし(A〜E列の仕事数が一致)")
            elif s["spillover"] == "drop":
                log("※ 先頭週の月・火の仕事は『載せない』設定のため出力に含みません。")
            if s["truncated"]:
                log("⚠ 月曜列に入り切らず省略: %d 件" % s["truncated"])
                ok_msg += "\n\n⚠ %d 件が月曜列に入り切らず省略されました。" % s["truncated"]
            log("✓ 完了")
            messagebox.showinfo("完了", ok_msg)
            try:
                os.startfile(os.path.dirname(s["output"]))
            except Exception:
                pass
        except PermissionError:
            messagebox.showerror("エラー", "ファイルを保存できません。\n出力先のExcelを閉じてから再実行してください。")
            log("保存に失敗(ファイルが開かれている可能性)")
        except Exception as e:
            log("エラー: %s" % e)
            log(traceback.format_exc())
            messagebox.showerror("エラー", "処理中にエラーが発生しました:\n%s" % e)
        finally:
            btn.config(state="normal")

    btn = ttk.Button(root, text="実行（2日前倒し）", command=do_run)
    btn.pack(pady=8)

    root.mainloop()

# ----------------------------------------------------------------------------
# エントリポイント(このファイル単体 = 予備のtkinter UI / CLI)
#   本体のハイブリッド版(ブラウザ画面)は koutei_app.py を起動してください。
# ----------------------------------------------------------------------------
def main():
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    if args and args[0].lower().endswith((".xlsx", ".xlsm")):
        # ファイルをドラッグ&ドロップ / コマンドライン指定 → その場で処理
        s = process(args[0], log=print)
        print("出力:", s["output"])
        return
    run_gui()  # 既定: 簡易UI(tkinter)


if __name__ == "__main__":
    main()

